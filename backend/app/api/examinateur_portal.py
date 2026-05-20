"""
Portail examinateur : accès via code_acces unique, saisie de notes.
"""
from typing import List, Optional
from datetime import date as Date

from fastapi import APIRouter, Depends, HTTPException, Security
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from jose import JWTError
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.auth import create_examinateur_token, decode_examinateur_token
from app.db.deps import get_db
from app.models.examinateur import Examinateur
from app.models.epreuve import Epreuve
from app.models.demi_journee import DemiJournee
from app.models.candidat import Candidat
from app.models.note import Note
from app.models.salle import Salle
from app.models.planche import Planche

router = APIRouter(prefix="/examinateur", tags=["examinateur-portal"])

_bearer = HTTPBearer()


# ── Guard ─────────────────────────────────────────────────────────────────────

def get_current_examinateur(
    credentials: HTTPAuthorizationCredentials = Security(_bearer),
    db: Session = Depends(get_db),
) -> Examinateur:
    try:
        ex_id = decode_examinateur_token(credentials.credentials)
    except JWTError:
        raise HTTPException(status_code=401, detail="Token invalide ou expiré")
    ex = db.get(Examinateur, ex_id)
    if not ex:
        raise HTTPException(status_code=401, detail="Examinateur introuvable")
    return ex


# ── Schemas ───────────────────────────────────────────────────────────────────

class LoginIn(BaseModel):
    code_acces: str


class LoginOut(BaseModel):
    access_token: str
    token_type: str = "bearer"


class ExaminateurMe(BaseModel):
    id: int
    nom: str
    prenom: str
    email: str
    matieres: List[str]

    class Config:
        from_attributes = True


class EpreuveExaminateur(BaseModel):
    id: int
    date: Date
    matiere: str
    heure_debut: str
    heure_fin: str
    preparation_minutes: Optional[int]
    candidat_id: Optional[int]
    candidat_nom: Optional[str]
    candidat_prenom: Optional[str]
    note_valeur: Optional[float]
    note_statut: Optional[str]
    salle_intitule: Optional[str]
    salle_preparation_intitule: Optional[str]
    planche_nom: Optional[str]
    conflit_etablissement: bool = False


class NoterIn(BaseModel):
    valeur: float


class NoterOut(BaseModel):
    note_id: int
    valeur: float
    statut: str


class CodePerduIn(BaseModel):
    email: str


# ── Routes ────────────────────────────────────────────────────────────────────

@router.post("/code-perdu")
def code_perdu(body: CodePerduIn, db: Session = Depends(get_db)):
    """Envoie le code d'accès par email si l'adresse est connue (réponse neutre)."""
    ex = db.query(Examinateur).filter(
        Examinateur.email == body.email.strip().lower()
    ).first()
    if ex:
        pass  # TODO: envoyer email avec ex.code_acces
    return {"sent": True}


@router.post("/login", response_model=LoginOut)
def login(body: LoginIn, db: Session = Depends(get_db)):
    ex = db.query(Examinateur).filter_by(code_acces=body.code_acces.strip().upper()).first()
    if not ex:
        raise HTTPException(status_code=401, detail="Code d'accès invalide")
    token = create_examinateur_token(ex.id)
    return LoginOut(access_token=token)


@router.get("/me", response_model=ExaminateurMe)
def me(ex: Examinateur = Depends(get_current_examinateur)):
    return ExaminateurMe(
        id=ex.id,
        nom=ex.nom,
        prenom=ex.prenom,
        email=ex.email,
        matieres=ex.matieres,
    )


@router.get("/me/epreuves", response_model=List[EpreuveExaminateur])
def mes_epreuves(
    ex: Examinateur = Depends(get_current_examinateur),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(Epreuve, DemiJournee)
        .join(DemiJournee, Epreuve.demi_journee_id == DemiJournee.id)
        .filter(Epreuve.examinateur_id == ex.id)
        .order_by(DemiJournee.date, Epreuve.heure_debut)
        .all()
    )

    result = []
    for epreuve, dj in rows:
        candidat = db.get(Candidat, epreuve.candidat_id) if epreuve.candidat_id else None
        note = None
        if candidat:
            note = (
                db.query(Note)
                .filter_by(candidat_id=candidat.id, matiere=epreuve.matiere)
                .first()
            )
        salle = db.get(Salle, epreuve.salle_id) if epreuve.salle_id else None
        salle_prep = db.get(Salle, epreuve.salle_preparation_id) if epreuve.salle_preparation_id else None
        planche = db.get(Planche, epreuve.planche_id) if epreuve.planche_id else None
        conflit = bool(
            candidat and ex.code_uai and candidat.code_uai
            and ex.code_uai.strip().upper() == candidat.code_uai.strip().upper()
        )
        result.append(EpreuveExaminateur(
            id=epreuve.id,
            date=dj.date,
            matiere=epreuve.matiere,
            heure_debut=str(epreuve.heure_debut)[:5],
            heure_fin=str(epreuve.heure_fin)[:5],
            preparation_minutes=epreuve.preparation_minutes,
            candidat_id=candidat.id if candidat else None,
            candidat_nom=candidat.nom if candidat else None,
            candidat_prenom=candidat.prenom if candidat else None,
            note_valeur=note.valeur if note else None,
            note_statut=note.statut if note else None,
            salle_intitule=salle.intitule if salle else None,
            salle_preparation_intitule=salle_prep.intitule if salle_prep else None,
            planche_nom=planche.nom if planche else None,
            conflit_etablissement=conflit,
        ))
    return result


@router.get("/me/epreuves/{epreuve_id}/planche")
def voir_planche(
    epreuve_id: int,
    ex: Examinateur = Depends(get_current_examinateur),
    db: Session = Depends(get_db),
):
    """Retourne le PDF de la planche assignée à une épreuve de l'examinateur."""
    import io
    from fastapi.responses import StreamingResponse
    epreuve = db.get(Epreuve, epreuve_id)
    if not epreuve or (epreuve.examinateur_id != ex.id and epreuve.examinateur2_id != ex.id):
        raise HTTPException(status_code=404, detail="Épreuve introuvable")
    if not epreuve.planche_id:
        raise HTTPException(status_code=404, detail="Aucun sujet assigné à cette épreuve")
    planche = db.get(Planche, epreuve.planche_id)
    if not planche or not planche.fichier_data:
        raise HTTPException(status_code=404, detail="Fichier introuvable")
    return StreamingResponse(
        io.BytesIO(planche.fichier_data),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{planche.nom}.pdf"'},
    )


@router.get("/me/epreuves/export")
def export_planning(
    ex: Examinateur = Depends(get_current_examinateur),
    db: Session = Depends(get_db),
):
    """Export Excel du planning + notes de l'examinateur."""
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import Response as _Response

    epreuves = mes_epreuves(ex=ex, db=db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planning"

    headers = ["Date", "Matière", "Heure prépa", "Heure passage", "Candidat", "Salle prépa", "Salle", "Sujet", "Note /20", "Statut note", "⚠ Conflit établ."]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B21B6")
        cell.alignment = Alignment(horizontal="center")

    JOURS = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
    MOIS = ["", "jan", "fév", "mar", "avr", "mai", "jun", "jul", "aoû", "sep", "oct", "nov", "déc"]

    for ep in epreuves:
        heure_prep = ""
        if ep.preparation_minutes:
            from datetime import datetime as _dt, timedelta
            h, m = map(int, ep.heure_debut.split(":"))
            prep_min = h * 60 + m - ep.preparation_minutes
            prep_min = ((prep_min % 1440) + 1440) % 1440
            heure_prep = f"{prep_min // 60:02d}:{prep_min % 60:02d}"

        d = ep.date
        date_str = f"{JOURS[d.weekday()]} {d.day} {MOIS[d.month]} {d.year}"
        candidat_str = f"{ep.candidat_prenom or ''} {ep.candidat_nom or ''}".strip() if ep.candidat_id else ""
        ws.append([
            date_str, ep.matiere, heure_prep, ep.heure_debut,
            candidat_str,
            ep.salle_preparation_intitule or "",
            ep.salle_intitule or "",
            ep.planche_nom or "",
            ep.note_valeur if ep.note_valeur is not None else "",
            ep.note_statut or "",
            "OUI" if ep.conflit_etablissement else "",
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = _io.BytesIO()
    wb.save(buf)
    nom = f"{ex.nom}_{ex.prenom}_planning".replace(" ", "_")
    return _Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nom}.xlsx"'},
    )


@router.post("/me/epreuves/{epreuve_id}/noter", response_model=NoterOut)
def noter(
    epreuve_id: int,
    body: NoterIn,
    ex: Examinateur = Depends(get_current_examinateur),
    db: Session = Depends(get_db),
):
    epreuve = db.get(Epreuve, epreuve_id)
    if not epreuve or epreuve.examinateur_id != ex.id:
        raise HTTPException(status_code=404, detail="Épreuve introuvable")
    if not epreuve.candidat_id:
        raise HTTPException(status_code=400, detail="Aucun candidat assigné à cette épreuve")
    if not (0 <= body.valeur <= 20):
        raise HTTPException(status_code=422, detail="La note doit être entre 0 et 20")

    note = (
        db.query(Note)
        .filter_by(candidat_id=epreuve.candidat_id, matiere=epreuve.matiere)
        .first()
    )
    if note:
        note.valeur = body.valeur
    else:
        note = Note(
            candidat_id=epreuve.candidat_id,
            matiere=epreuve.matiere,
            valeur=body.valeur,
            statut="BROUILLON",
        )
        db.add(note)
    db.commit()
    db.refresh(note)
    return NoterOut(note_id=note.id, valeur=note.valeur, statut=note.statut)
