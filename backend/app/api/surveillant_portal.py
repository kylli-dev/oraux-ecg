"""
Portail surveillant : accès via code_acces, consultation planning, marquage absences.
"""
import io
from datetime import date as Date
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Security
from fastapi.responses import Response, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from jose import JWTError
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.auth import create_surveillant_token, decode_surveillant_token
from app.db.deps import get_db
from app.models.surveillant import Surveillant
from app.models.epreuve import Epreuve
from app.models.demi_journee import DemiJournee
from app.models.candidat import Candidat
from app.models.salle import Salle
from app.models.planche import Planche
from app.models.examinateur import Examinateur

router = APIRouter(prefix="/surveillant", tags=["surveillant-portal"])

_bearer = HTTPBearer()


# ── Guard ─────────────────────────────────────────────────────────────────────

def get_current_surveillant(
    credentials: HTTPAuthorizationCredentials = Security(_bearer),
    db: Session = Depends(get_db),
) -> Surveillant:
    try:
        sid = decode_surveillant_token(credentials.credentials)
    except JWTError:
        raise HTTPException(status_code=401, detail="Token invalide ou expiré")
    s = db.get(Surveillant, sid)
    if not s:
        raise HTTPException(status_code=401, detail="Surveillant introuvable")
    return s


# ── Schemas ───────────────────────────────────────────────────────────────────

class LoginIn(BaseModel):
    code_acces: str


class LoginOut(BaseModel):
    access_token: str


class SurveillantMe(BaseModel):
    id: int
    nom: str
    prenom: str
    email: str


class EpreuveSurveillant(BaseModel):
    id: int
    date: Date
    matiere: str
    heure_debut: str
    heure_fin: str
    preparation_minutes: Optional[int] = None
    statut: str
    candidat_id: Optional[int] = None
    candidat_nom: Optional[str] = None
    candidat_prenom: Optional[str] = None
    salle_intitule: Optional[str] = None
    salle_preparation_intitule: Optional[str] = None
    planche_nom: Optional[str] = None
    examinateur_nom: Optional[str] = None
    examinateur_prenom: Optional[str] = None


# ── Routes ────────────────────────────────────────────────────────────────────

@router.post("/login", response_model=LoginOut)
def login(body: LoginIn, db: Session = Depends(get_db)):
    from sqlalchemy import func as _func
    s = db.query(Surveillant).filter(
        _func.upper(Surveillant.code_acces) == body.code_acces.strip().upper()
    ).first()
    if not s:
        raise HTTPException(status_code=401, detail="Code d'accès invalide")
    return LoginOut(access_token=create_surveillant_token(s.id))


@router.get("/me", response_model=SurveillantMe)
def me(s: Surveillant = Depends(get_current_surveillant)):
    return SurveillantMe(id=s.id, nom=s.nom, prenom=s.prenom, email=s.email)


@router.get("/me/epreuves", response_model=List[EpreuveSurveillant])
def mes_epreuves(
    s: Surveillant = Depends(get_current_surveillant),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(Epreuve, DemiJournee)
        .join(DemiJournee, Epreuve.demi_journee_id == DemiJournee.id)
        .filter(Epreuve.surveillant_id == s.id)
        .order_by(DemiJournee.date, Epreuve.heure_debut)
        .all()
    )

    result = []
    for ep, dj in rows:
        candidat = db.get(Candidat, ep.candidat_id) if ep.candidat_id else None
        salle = db.get(Salle, ep.salle_id) if ep.salle_id else None
        salle_prep = db.get(Salle, ep.salle_preparation_id) if ep.salle_preparation_id else None
        planche = db.get(Planche, ep.planche_id) if ep.planche_id else None
        exam = db.get(Examinateur, ep.examinateur_id) if ep.examinateur_id else None
        result.append(EpreuveSurveillant(
            id=ep.id,
            date=dj.date,
            matiere=ep.matiere,
            heure_debut=str(ep.heure_debut)[:5],
            heure_fin=str(ep.heure_fin)[:5],
            preparation_minutes=ep.preparation_minutes,
            statut=ep.statut,
            candidat_id=candidat.id if candidat else None,
            candidat_nom=candidat.nom if candidat else None,
            candidat_prenom=candidat.prenom if candidat else None,
            salle_intitule=salle.intitule if salle else None,
            salle_preparation_intitule=salle_prep.intitule if salle_prep else None,
            planche_nom=planche.nom if planche else None,
            examinateur_nom=exam.nom if exam else None,
            examinateur_prenom=exam.prenom if exam else None,
        ))
    return result


@router.post("/me/epreuves/{epreuve_id}/marquer-absent", status_code=204)
def marquer_absent(
    epreuve_id: int,
    s: Surveillant = Depends(get_current_surveillant),
    db: Session = Depends(get_db),
):
    ep = db.get(Epreuve, epreuve_id)
    if not ep or ep.surveillant_id != s.id:
        raise HTTPException(status_code=404, detail="Épreuve introuvable")
    if not ep.candidat_id:
        raise HTTPException(status_code=400, detail="Aucun candidat assigné")
    if ep.statut != "ATTRIBUEE":
        raise HTTPException(status_code=400, detail=f"Statut actuel : {ep.statut}")
    ep.statut = "ABSENT"
    db.commit()


@router.delete("/me/epreuves/{epreuve_id}/marquer-absent", status_code=204)
def annuler_absent(
    epreuve_id: int,
    s: Surveillant = Depends(get_current_surveillant),
    db: Session = Depends(get_db),
):
    ep = db.get(Epreuve, epreuve_id)
    if not ep or ep.surveillant_id != s.id:
        raise HTTPException(status_code=404, detail="Épreuve introuvable")
    if ep.statut != "ABSENT":
        raise HTTPException(status_code=400, detail="L'épreuve n'est pas marquée absente")
    ep.statut = "ATTRIBUEE"
    db.commit()


@router.get("/me/epreuves/export")
def export_planning(
    s: Surveillant = Depends(get_current_surveillant),
    db: Session = Depends(get_db),
):
    """Export Excel du planning de surveillance."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    epreuves = mes_epreuves(s=s, db=db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planning surveillance"

    headers = ["Date", "Matière", "Heure prépa", "Heure passage", "Candidat", "Salle prépa", "Salle", "Examinateur", "Sujet", "Statut"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C62828")
        cell.alignment = Alignment(horizontal="center")

    JOURS = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
    MOIS = ["", "jan", "fév", "mar", "avr", "mai", "jun", "jul", "aoû", "sep", "oct", "nov", "déc"]

    for ep in epreuves:
        heure_prep = ""
        if ep.preparation_minutes:
            h, m = map(int, ep.heure_debut.split(":"))
            prep_min = ((h * 60 + m - ep.preparation_minutes) % 1440 + 1440) % 1440
            heure_prep = f"{prep_min // 60:02d}:{prep_min % 60:02d}"

        d = ep.date
        date_str = f"{JOURS[d.weekday()]} {d.day} {MOIS[d.month]} {d.year}"
        candidat_str = f"{ep.candidat_prenom or ''} {ep.candidat_nom or ''}".strip() if ep.candidat_id else "—"
        exam_str = f"{ep.examinateur_prenom or ''} {ep.examinateur_nom or ''}".strip() if ep.examinateur_nom else "—"
        ws.append([
            date_str, ep.matiere, heure_prep, ep.heure_debut,
            candidat_str, ep.salle_preparation_intitule or "—", ep.salle_intitule or "—",
            exam_str, ep.planche_nom or "—", ep.statut,
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    nom = f"{s.nom}_{s.prenom}_surveillance".replace(" ", "_")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nom}.xlsx"'},
    )
