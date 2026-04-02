"""
Gestion admin des notes : consultation, publication individuelle ou en masse.
"""
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.admin_guard import require_admin
from app.db.deps import get_db
from app.models.note import Note
from app.models.candidat import Candidat

router = APIRouter(
    prefix="/admin/notes",
    tags=["notes"],
    dependencies=[Depends(require_admin)],
)


# ── Schemas ───────────────────────────────────────────────────────────────────

class NoteOut(BaseModel):
    id: int
    candidat_id: int
    candidat_nom: str
    candidat_prenom: str
    matiere: str
    valeur: Optional[float]
    note_harmonisee: Optional[float]
    commentaire: Optional[str]
    statut: str

    class Config:
        from_attributes = True


class PublierBulkIn(BaseModel):
    note_ids: List[int]


class HarmoniserIn(BaseModel):
    note_harmonisee: Optional[float] = None
    commentaire: Optional[str] = None


class SaisieEpreuveOut(BaseModel):
    epreuve_id: int
    date: str
    matiere: str
    heure_debut: str
    heure_fin: str
    preparation_minutes: Optional[int]
    candidat_id: Optional[int]
    candidat_nom: Optional[str]
    candidat_prenom: Optional[str]
    note_id: Optional[int]
    valeur: Optional[float]
    commentaire: Optional[str]
    statut: Optional[str]


class SaisirNoteIn(BaseModel):
    epreuve_id: int
    valeur: Optional[float] = None
    commentaire: Optional[str] = None


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get("/", response_model=List[NoteOut])
def list_notes(
    statut: Optional[str] = None,
    candidat_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    q = db.query(Note, Candidat).join(Candidat, Note.candidat_id == Candidat.id)
    if statut:
        q = q.filter(Note.statut == statut.upper())
    if candidat_id:
        q = q.filter(Note.candidat_id == candidat_id)
    q = q.order_by(Candidat.nom, Candidat.prenom, Note.matiere)

    return [
        NoteOut(
            id=note.id,
            candidat_id=note.candidat_id,
            candidat_nom=candidat.nom,
            candidat_prenom=candidat.prenom,
            matiere=note.matiere,
            valeur=note.valeur,
            note_harmonisee=note.note_harmonisee,
            commentaire=note.commentaire,
            statut=note.statut,
        )
        for note, candidat in q.all()
    ]


@router.post("/{note_id}/publier", response_model=NoteOut)
def publier_note(note_id: int, db: Session = Depends(get_db)):
    note = db.get(Note, note_id)
    if not note:
        raise HTTPException(status_code=404, detail="Note introuvable")
    note.statut = "PUBLIE"
    note.published_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(note)
    candidat = db.get(Candidat, note.candidat_id)
    return NoteOut(
        id=note.id,
        candidat_id=note.candidat_id,
        candidat_nom=candidat.nom,
        candidat_prenom=candidat.prenom,
        matiere=note.matiere,
        valeur=note.valeur,
        note_harmonisee=note.note_harmonisee,
        commentaire=note.commentaire,
        statut=note.statut,
    )


@router.post("/publier-bulk")
def publier_bulk(body: PublierBulkIn, db: Session = Depends(get_db)):
    """Publie plusieurs notes d'un coup."""
    count = 0
    now = datetime.now(timezone.utc)
    for note_id in body.note_ids:
        note = db.get(Note, note_id)
        if note and note.statut == "BROUILLON":
            note.statut = "PUBLIE"
            note.published_at = now
            count += 1
    db.commit()
    return {"publiees": count}


@router.post("/publier-tout")
def publier_tout(db: Session = Depends(get_db)):
    """Publie toutes les notes en statut BROUILLON."""
    notes = db.query(Note).filter_by(statut="BROUILLON").all()
    now = datetime.now(timezone.utc)
    for note in notes:
        note.statut = "PUBLIE"
        note.published_at = now
    db.commit()
    return {"publiees": len(notes)}


# ── Harmonisation ────────────────────────────────────────────────────────────

@router.patch("/{note_id}/harmoniser", response_model=NoteOut)
def harmoniser_note(note_id: int, body: HarmoniserIn, db: Session = Depends(get_db)):
    """Met à jour la note harmonisée (et/ou le commentaire) d'une note."""
    note = db.get(Note, note_id)
    if not note:
        raise HTTPException(status_code=404, detail="Note introuvable")
    if body.note_harmonisee is not None:
        note.note_harmonisee = body.note_harmonisee
    if body.commentaire is not None:
        note.commentaire = body.commentaire
    db.commit()
    db.refresh(note)
    candidat = db.get(Candidat, note.candidat_id)
    return NoteOut(
        id=note.id,
        candidat_id=note.candidat_id,
        candidat_nom=candidat.nom,
        candidat_prenom=candidat.prenom,
        matiere=note.matiere,
        valeur=note.valeur,
        note_harmonisee=note.note_harmonisee,
        commentaire=note.commentaire,
        statut=note.statut,
    )


# ── Saisie admin pour un examinateur ────────────────────────────────────────

@router.get("/saisie", response_model=List[SaisieEpreuveOut])
def get_saisie(
    examinateur_id: int,
    planning_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """
    Retourne toutes les épreuves d'un examinateur avec les notes actuelles.
    Permet au service des admissions de saisir les notes à la place de l'examinateur.
    """
    from app.models.epreuve import Epreuve
    from app.models.demi_journee import DemiJournee

    q = (
        db.query(Epreuve, DemiJournee)
        .join(DemiJournee, Epreuve.demi_journee_id == DemiJournee.id)
        .filter(Epreuve.examinateur_id == examinateur_id)
    )
    if planning_id:
        q = q.filter(DemiJournee.planning_id == planning_id)
    rows = q.order_by(DemiJournee.date, Epreuve.heure_debut).all()

    result = []
    for ep, dj in rows:
        candidat = db.get(Candidat, ep.candidat_id) if ep.candidat_id else None
        note = None
        if candidat:
            note = db.query(Note).filter_by(
                candidat_id=candidat.id, matiere=ep.matiere
            ).first()
        result.append(SaisieEpreuveOut(
            epreuve_id=ep.id,
            date=str(dj.date),
            matiere=ep.matiere,
            heure_debut=str(ep.heure_debut)[:5],
            heure_fin=str(ep.heure_fin)[:5],
            preparation_minutes=ep.preparation_minutes,
            candidat_id=candidat.id if candidat else None,
            candidat_nom=candidat.nom if candidat else None,
            candidat_prenom=candidat.prenom if candidat else None,
            note_id=note.id if note else None,
            valeur=note.valeur if note else None,
            commentaire=note.commentaire if note else None,
            statut=note.statut if note else None,
        ))
    return result


@router.post("/saisir", response_model=NoteOut)
def saisir_note(body: SaisirNoteIn, db: Session = Depends(get_db)):
    """
    Crée ou met à jour une note pour une épreuve (saisie admin).
    La valeur peut être None pour effacer la note.
    """
    from app.models.epreuve import Epreuve

    ep = db.get(Epreuve, body.epreuve_id)
    if not ep:
        raise HTTPException(status_code=404, detail="Épreuve introuvable")
    if not ep.candidat_id:
        raise HTTPException(status_code=400, detail="Aucun candidat assigné à cette épreuve")
    if body.valeur is not None and not (0 <= body.valeur <= 20):
        raise HTTPException(status_code=422, detail="La note doit être entre 0 et 20")

    note = db.query(Note).filter_by(
        candidat_id=ep.candidat_id, matiere=ep.matiere
    ).first()

    if note:
        if body.valeur is not None:
            note.valeur = body.valeur
        if body.commentaire is not None:
            note.commentaire = body.commentaire
    else:
        note = Note(
            candidat_id=ep.candidat_id,
            matiere=ep.matiere,
            valeur=body.valeur,
            commentaire=body.commentaire,
            statut="BROUILLON",
        )
        db.add(note)

    db.commit()
    db.refresh(note)
    candidat = db.get(Candidat, note.candidat_id)
    return NoteOut(
        id=note.id,
        candidat_id=note.candidat_id,
        candidat_nom=candidat.nom,
        candidat_prenom=candidat.prenom or "",
        matiere=note.matiere,
        valeur=note.valeur,
        note_harmonisee=note.note_harmonisee,
        commentaire=note.commentaire,
        statut=note.statut,
    )


# ── Tableau pivot ────────────────────────────────────────────────────────────

def _build_tableau(planning_id: int, db: Session) -> Dict[str, Any]:
    """
    Retourne la structure pivot:
    { matieres: [...], candidats: [{id, nom, prenom, code_candidat, date, notes: {matiere: {...}}}] }
    """
    from app.models.epreuve import Epreuve
    from app.models.demi_journee import DemiJournee

    # Candidats inscrits dans ce planning (via leurs épreuves ATTRIBUEE)
    rows = (
        db.query(Epreuve, DemiJournee)
        .join(DemiJournee, Epreuve.demi_journee_id == DemiJournee.id)
        .filter(
            DemiJournee.planning_id == planning_id,
            Epreuve.candidat_id.isnot(None),
        )
        .order_by(DemiJournee.date, Epreuve.heure_debut)
        .all()
    )

    # Collect candidat_ids, their date (first épreuve), matieres
    candidat_dates: Dict[int, str] = {}
    matieres_set: set = set()
    for ep, dj in rows:
        if ep.candidat_id not in candidat_dates:
            candidat_dates[ep.candidat_id] = str(dj.date)
        matieres_set.add(ep.matiere)

    matieres = sorted(matieres_set)

    # Get all notes for these candidats
    notes_by_cand: Dict[int, Dict[str, NoteOut]] = {}
    if candidat_dates:
        note_rows = (
            db.query(Note, Candidat)
            .join(Candidat, Note.candidat_id == Candidat.id)
            .filter(Note.candidat_id.in_(list(candidat_dates.keys())))
            .all()
        )
        for note, candidat in note_rows:
            notes_by_cand.setdefault(note.candidat_id, {})[note.matiere] = {
                "note_id": note.id,
                "valeur": note.valeur,
                "note_harmonisee": note.note_harmonisee,
                "ecart": (
                    round(note.note_harmonisee - note.valeur, 2)
                    if note.note_harmonisee is not None and note.valeur is not None
                    else None
                ),
                "commentaire": note.commentaire,
                "statut": note.statut,
            }

    # Build candidats list sorted by date then nom
    seen: set = set()
    candidats_out = []
    for ep, dj in rows:
        cid = ep.candidat_id
        if cid in seen:
            continue
        seen.add(cid)
        c = db.get(Candidat, cid)
        if not c:
            continue
        candidats_out.append({
            "id": c.id,
            "nom": c.nom,
            "prenom": c.prenom or "",
            "code_candidat": c.code_candidat,
            "handicape": c.handicape,
            "date": candidat_dates.get(cid, ""),
            "notes": notes_by_cand.get(cid, {}),
        })

    return {"matieres": matieres, "candidats": candidats_out}


@router.get("/tableau")
def get_tableau(planning_id: int, db: Session = Depends(get_db)):
    """Tableau pivot candidats × matières avec notes brutes et harmonisées."""
    return _build_tableau(planning_id, db)


@router.get("/tableau/export")
def export_tableau(planning_id: int, db: Session = Depends(get_db)):
    """Export Excel du tableau pivot."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    tableau = _build_tableau(planning_id, db)
    matieres: List[str] = tableau["matieres"]
    candidats = tableau["candidats"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notes"

    # ── Styles ──────────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1A237E")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    sub_fill = PatternFill("solid", fgColor="E8EAF6")
    sub_font = Font(bold=True, size=8, color="1A237E")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ── En-têtes ─────────────────────────────────────────────────────────────────
    # Row 1: fixed columns + matière headers (merged across 3 sub-cols: Note, Harm., Écart)
    fixed_cols = ["Nom", "Prénom", "N° Candidat", "Handicap", "Date"]
    n_fixed = len(fixed_cols)

    # Row 1
    for col_i, label in enumerate(fixed_cols, start=1):
        c = ws.cell(row=1, column=col_i, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin
        ws.merge_cells(start_row=1, start_column=col_i, end_row=2, end_column=col_i)

    for m_i, mat in enumerate(matieres):
        col_start = n_fixed + 1 + m_i * 3
        c = ws.cell(row=1, column=col_start, value=mat)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin
        ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start + 2)

    # Row 2: sub-headers per matière
    sub_labels = ["Note", "Harm.", "Écart"]
    for m_i in range(len(matieres)):
        for s_i, sl in enumerate(sub_labels):
            col = n_fixed + 1 + m_i * 3 + s_i
            c = ws.cell(row=2, column=col, value=sl)
            c.font = sub_font
            c.fill = sub_fill
            c.alignment = center
            c.border = thin

    # ── Données ──────────────────────────────────────────────────────────────────
    for r_i, cand in enumerate(candidats, start=3):
        fixed_values = [
            cand["nom"].upper(),
            cand["prenom"],
            cand["code_candidat"] or "",
            "Oui" if cand["handicape"] else "",
            cand["date"],
        ]
        for col_i, val in enumerate(fixed_values, start=1):
            c = ws.cell(row=r_i, column=col_i, value=val)
            c.alignment = left
            c.border = thin

        for m_i, mat in enumerate(matieres):
            note_data = cand["notes"].get(mat, {})
            vals = [
                note_data.get("valeur"),
                note_data.get("note_harmonisee"),
                note_data.get("ecart"),
            ]
            for s_i, val in enumerate(vals):
                col = n_fixed + 1 + m_i * 3 + s_i
                c = ws.cell(row=r_i, column=col, value=val)
                c.alignment = center
                c.border = thin
                if val is not None and isinstance(val, (int, float)):
                    c.number_format = "0.00"

    # ── Largeurs colonnes ────────────────────────────────────────────────────────
    col_widths = [18, 14, 12, 9, 12] + [8, 8, 7] * len(matieres)
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="notes_planning_{planning_id}.xlsx"'},
    )
