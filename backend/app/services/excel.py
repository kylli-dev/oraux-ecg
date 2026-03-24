"""
Import / Export Excel pour les plannings, candidats et examinateurs.

Export : toutes les épreuves d'un planning (avec candidat si assigné).
Import : création d'épreuves depuis un fichier Excel.
         Format attendu (ligne 1 = en-têtes ignorées) :
         A: date (YYYY-MM-DD)  B: heure_debut (HH:MM)  C: heure_fin (HH:MM)
         D: matiere            E: statut (optionnel, défaut LIBRE)
"""
import io
import json
import re
import secrets
import unicodedata
from datetime import date as date_type, time as time_type

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from app.models.candidat import Candidat
from app.models.demi_journee import DemiJournee
from app.models.epreuve import Epreuve
from app.models.examinateur import Examinateur
from app.models.planning import Planning

RED_HEX = "C62828"
GREY_HEX = "F5F5F5"

ALLOWED_STATUT = {"CREE", "LIBRE", "ATTRIBUEE", "EN_EVALUATION", "FINALISEE", "ANNULEE"}


# ── Export ─────────────────────────────────────────────────────────────────────

def export_planning(db: Session, planning_id: int) -> bytes:
    planning = db.get(Planning, planning_id)
    if not planning:
        raise ValueError("Planning not found")

    wb = openpyxl.Workbook()

    # ── Sheet 1 : Planning info ────────────────────────────────────────────────
    ws_info = wb.active
    ws_info.title = "Planning"
    _header_row(ws_info, ["Champ", "Valeur"])
    for row in [
        ("ID", planning.id),
        ("Nom", planning.nom),
        ("Date début", str(planning.date_debut)),
        ("Date fin", str(planning.date_fin)),
        ("Statut", planning.statut),
        ("Ouverture inscriptions", str(planning.date_ouverture_inscriptions)),
        ("Fermeture inscriptions", str(planning.date_fermeture_inscriptions)),
    ]:
        ws_info.append(row)

    # ── Sheet 2 : Épreuves ────────────────────────────────────────────────────
    ws_ep = wb.create_sheet("Epreuves")
    headers = [
        "Date", "Demi-journée", "Heure début", "Heure fin",
        "Matière", "Statut", "Candidat — Nom", "Candidat — Prénom", "Email",
    ]
    _header_row(ws_ep, headers)

    djs = (
        db.query(DemiJournee)
        .filter_by(planning_id=planning_id)
        .order_by(DemiJournee.date, DemiJournee.heure_debut)
        .all()
    )
    for dj in djs:
        epreuves = (
            db.query(Epreuve)
            .filter_by(demi_journee_id=dj.id)
            .order_by(Epreuve.heure_debut)
            .all()
        )
        for e in epreuves:
            cand = db.get(Candidat, e.candidat_id) if e.candidat_id else None
            ws_ep.append([
                str(dj.date),
                dj.type,
                str(e.heure_debut)[:5],
                str(e.heure_fin)[:5],
                e.matiere,
                e.statut,
                cand.nom if cand else "",
                cand.prenom if cand else "",
                cand.email if cand else "",
            ])

    _autosize(ws_ep)

    # ── Sheet 3 : Candidats ────────────────────────────────────────────────────
    ws_cand = wb.create_sheet("Candidats")
    _header_row(ws_cand, ["ID", "Nom", "Prénom", "Email", "Code accès", "Statut"])
    candidats = db.query(Candidat).filter_by(planning_id=planning_id).all()
    for c in candidats:
        ws_cand.append([c.id, c.nom, c.prenom, c.email, c.code_acces, c.statut])
    _autosize(ws_cand)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Template vierge ────────────────────────────────────────────────────────────

def export_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Epreuves"
    _header_row(ws, ["date", "heure_debut", "heure_fin", "matiere", "statut"])

    # Lignes d'exemple
    ws.append(["2026-06-15", "08:30", "09:00", "Maths", "LIBRE"])
    ws.append(["2026-06-15", "09:05", "09:35", "Anglais", "LIBRE"])
    ws.append(["2026-06-15", "13:30", "14:00", "Economie", "LIBRE"])

    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Import ─────────────────────────────────────────────────────────────────────

def import_epreuves(db: Session, planning_id: int, file_bytes: bytes) -> dict:
    planning = db.get(Planning, planning_id)
    if not planning:
        raise ValueError("Planning not found")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    created = 0
    errors: list[str] = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            d = _parse_date(row[0])
            hdebut = _parse_time(row[1])
            hfin = _parse_time(row[2])
            matiere = str(row[3]).strip() if row[3] else None
            statut = str(row[4]).strip().upper() if len(row) > 4 and row[4] else "LIBRE"

            if not all([d, hdebut, hfin, matiere]):
                errors.append(f"Ligne {i} : données manquantes")
                continue

            if statut not in ALLOWED_STATUT:
                statut = "LIBRE"

            type_dj = "MATIN" if hdebut.hour < 12 else "APRES_MIDI"

            # Upsert demi-journée
            dj = (
                db.query(DemiJournee)
                .filter_by(planning_id=planning_id, date=d, type=type_dj)
                .first()
            )
            if dj is None:
                dj = DemiJournee(
                    planning_id=planning_id,
                    date=d,
                    type=type_dj,
                    heure_debut=hdebut,
                    heure_fin=hfin,
                )
                db.add(dj)
                db.flush()
            else:
                if hdebut < dj.heure_debut:
                    dj.heure_debut = hdebut
                if hfin > dj.heure_fin:
                    dj.heure_fin = hfin
                db.flush()

            e = Epreuve(
                demi_journee_id=dj.id,
                matiere=matiere,
                heure_debut=hdebut,
                heure_fin=hfin,
                statut=statut,
            )
            db.add(e)
            created += 1

        except Exception as ex:
            errors.append(f"Ligne {i} : {ex}")

    db.commit()
    return {"epreuves_created": created, "errors": errors}


# ── Import Candidats ───────────────────────────────────────────────────────────

def import_candidats(db: Session, planning_id: int, file_bytes: bytes) -> dict:
    """
    Importe des candidats depuis un Excel.
    Colonnes attendues (ligne 1 = en-têtes) :
      A: nom  B: prenom  C: email  D: login (optionnel)
      E: profil (HGG/ESH, optionnel)  F: code_uai (optionnel)
    Retourne la liste des créations avec mot de passe provisoire.
    """
    from app.core.auth import hash_password

    planning = db.get(Planning, planning_id)
    if not planning:
        raise ValueError("Planning not found")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    created = []
    errors: list[str] = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            nom = str(row[0]).strip() if row[0] else None
            prenom = str(row[1]).strip() if row[1] else None
            email = str(row[2]).strip() if row[2] else None
            login = str(row[3]).strip() if len(row) > 3 and row[3] else None
            profil = str(row[4]).strip().upper() if len(row) > 4 and row[4] else None
            code_uai = str(row[5]).strip() if len(row) > 5 and row[5] else None

            if not all([nom, prenom, email]):
                errors.append(f"Ligne {i} : nom, prénom et email obligatoires")
                continue

            if profil and profil not in ("HGG", "ESH"):
                profil = None

            # Générer un mot de passe provisoire
            plain_pwd = secrets.token_urlsafe(8)

            c = Candidat(
                planning_id=planning_id,
                nom=nom,
                prenom=prenom,
                email=email,
                login=login or email,
                password_hash=hash_password(plain_pwd, rounds=4),
                profil=profil,
                code_uai=code_uai,
            )
            db.add(c)
            db.flush()
            created.append({
                "id": c.id,
                "nom": nom,
                "prenom": prenom,
                "email": email,
                "login": c.login,
                "password_provisoire": plain_pwd,
            })
        except Exception as ex:
            errors.append(f"Ligne {i} : {ex}")

    db.commit()
    return {"created": len(created), "candidats": created, "errors": errors}


def export_template_candidats() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidats"
    _header_row(ws, ["nom", "prenom", "email", "login (optionnel)", "profil (HGG/ESH)", "code_uai (optionnel)"])
    ws.append(["DUPONT", "Alice", "alice.dupont@exemple.fr", "alice.dupont", "HGG", "0751234A"])
    ws.append(["MARTIN", "Bob", "bob.martin@exemple.fr", "", "ESH", ""])
    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Import Examinateurs ────────────────────────────────────────────────────────

def import_examinateurs(db: Session, planning_id: int, file_bytes: bytes) -> dict:
    """
    Importe des examinateurs depuis un Excel.
    Colonnes attendues :
      A: nom  B: prenom  C: email
      D: matieres (séparées par ';')  E: code_uai (optionnel)
    """
    planning = db.get(Planning, planning_id)
    if not planning:
        raise ValueError("Planning not found")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    created = []
    errors: list[str] = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            nom = str(row[0]).strip() if row[0] else None
            prenom = str(row[1]).strip() if row[1] else None
            email = str(row[2]).strip() if row[2] else None
            matieres_raw = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            code_uai = str(row[4]).strip() if len(row) > 4 and row[4] else None

            if not all([nom, prenom, email]):
                errors.append(f"Ligne {i} : nom, prénom et email obligatoires")
                continue

            matieres = [m.strip() for m in matieres_raw.split(";") if m.strip()]

            ex = Examinateur(
                planning_id=planning_id,
                nom=nom,
                prenom=prenom,
                email=email,
                matieres_json=json.dumps(matieres),
                code_uai=code_uai,
            )
            db.add(ex)
            db.flush()
            created.append({
                "id": ex.id,
                "nom": nom,
                "prenom": prenom,
                "email": email,
                "code_acces": ex.code_acces,
                "matieres": matieres,
            })
        except Exception as ex_err:
            errors.append(f"Ligne {i} : {ex_err}")

    db.commit()
    return {"created": len(created), "examinateurs": created, "errors": errors}


def export_template_examinateurs() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Examinateurs"
    _header_row(ws, ["nom", "prenom", "email", "matieres (séparées par ;)", "code_uai (optionnel)"])
    ws.append(["BERNARD", "Claire", "claire.bernard@lycee.fr", "Maths;Physique", "0751234A"])
    ws.append(["LEROY", "Paul", "paul.leroy@lycee.fr", "Anglais", ""])
    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Template candidats complet (18 champs) ────────────────────────────────────

CANDIDATS_COMPLET_HEADERS = [
    "CODE_CANDIDAT",
    "NUMERO_INE",
    "CIVILITE",
    "NOM",
    "PRENOM",
    "DATE_NAISSANCE",
    "QUALITE",
    "HANDICAPE",
    "CP",
    "VILLE",
    "LIBELLE_PAYS",
    "TEL_PORTABLE",
    "EMAIL",
    "CLASSE",
    "NUMERO_RNE",
    "ETABLISSEMENT",
    "VILLE_ETABLISSEMENT",
    "DEPARTEMENT_ETABLISSEMENT",
]


def export_template_candidats_complet() -> bytes:
    """Fichier modèle d'import candidats avec les 18 champs du concours."""
    wb = openpyxl.Workbook()

    # ── Feuille 1 : données ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Candidats"
    _header_row(ws, CANDIDATS_COMPLET_HEADERS)

    # Ligne d'exemple
    ws.append([
        "15617",                          # CODE_CANDIDAT
        "070164375KA",                    # NUMERO_INE
        "M.",                             # CIVILITE  (M. / Mme.)
        "IKSIL",                          # NOM
        "Raphaël",                        # PRENOM
        "30/03/2005",                     # DATE_NAISSANCE  (DD/MM/YYYY)
        "Non boursier",                   # QUALITE
        "Non",                            # HANDICAPE  (Oui / Non)
        "75015",                          # CP
        "Paris",                          # VILLE
        "France",                         # LIBELLE_PAYS
        "0638348849",                     # TEL_PORTABLE
        "raphaeliksil@gmail.com",         # EMAIL
        "ECG - Maths approfondies et HGG",  # CLASSE
        "0750658H",                       # NUMERO_RNE
        "Lycée Saint-Louis",              # ETABLISSEMENT
        "Paris",                          # VILLE_ETABLISSEMENT
        "75",                             # DEPARTEMENT_ETABLISSEMENT
    ])

    # Validation CIVILITE
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_civilite = DataValidation(type="list", formula1='"M.,Mme."', allow_blank=True)
    ws.add_data_validation(dv_civilite)
    dv_civilite.sqref = "C2:C5000"

    # Validation HANDICAPE
    dv_handicap = DataValidation(type="list", formula1='"Oui,Non"', allow_blank=True)
    ws.add_data_validation(dv_handicap)
    dv_handicap.sqref = "H2:H5000"

    _autosize(ws)

    # ── Feuille 2 : instructions ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("Champ", "Description", "Format", "Obligatoire"),
        ("CODE_CANDIDAT", "Code candidat attribué par le concours", "Texte libre", "Non"),
        ("NUMERO_INE", "Identifiant national élève (11 caractères)", "Alphanumérique", "Non"),
        ("CIVILITE", "Civilité", "M. ou Mme.", "Non"),
        ("NOM", "Nom de famille (majuscules recommandées)", "Texte", "Oui"),
        ("PRENOM", "Prénom", "Texte", "Oui"),
        ("DATE_NAISSANCE", "Date de naissance", "DD/MM/YYYY", "Non"),
        ("QUALITE", "Qualité / statut concours (ex : ADMISSIBLE)", "Texte", "Non"),
        ("HANDICAPE", "Aménagement handicap", "Oui ou Non", "Non"),
        ("CP", "Code postal du domicile", "Texte (5 chiffres)", "Non"),
        ("VILLE", "Ville du domicile", "Texte", "Non"),
        ("LIBELLE_PAYS", "Pays du domicile", "Texte", "Non"),
        ("TEL_PORTABLE", "Numéro de téléphone portable", "Texte", "Non"),
        ("EMAIL", "Adresse e-mail (utilisée comme login)", "Email valide", "Oui"),
        ("CLASSE", "Classe (ex : ECG1, ECG2)", "Texte", "Non"),
        ("NUMERO_RNE", "Numéro RNE / code UAI de l'établissement", "8 caractères", "Non"),
        ("ETABLISSEMENT", "Nom de l'établissement d'origine", "Texte", "Non"),
        ("VILLE_ETABLISSEMENT", "Ville de l'établissement", "Texte", "Non"),
        ("DEPARTEMENT_ETABLISSEMENT", "Département de l'établissement", "Texte", "Non"),
    ]
    for row in instructions:
        ws2.append(row)
    # Mettre en forme l'en-tête
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=RED_HEX)
        cell.alignment = Alignment(horizontal="center")
    _autosize(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Normalisation des en-têtes ────────────────────────────────────────────────

def _norm(s: str) -> str:
    """Normalise un en-tête : majuscules, sans accents, tout séparateur → _."""
    s = s.upper().strip()
    # Supprimer les accents
    s = "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )
    # Remplacer tout caractère non alphanumérique par _ (couvre °, /, -, espace, N°…)
    s = re.sub(r"[^A-Z0-9]+", "_", s)
    s = s.strip("_")
    return s


# Table d'alias : clé canonique → liste d'alias normalisés acceptés
_COLUMN_ALIASES: dict[str, list[str]] = {
    "NOM":                      ["NOM", "NOM_DE_FAMILLE", "LASTNAME", "NAME", "FAMILY_NAME"],
    "PRENOM":                   ["PRENOM", "FIRSTNAME", "FIRST_NAME", "PRENOMS"],
    "EMAIL":                    ["EMAIL", "E_MAIL", "MAIL", "EMAIL_PERSO", "E_MAIL_PERSO",
                                 "COURRIEL", "ADRESSE_MAIL", "ADRESSE_EMAIL", "ADRESSE_E_MAIL"],
    "CODE_CANDIDAT":            ["CODE_CANDIDAT", "N_CANDIDAT", "NO_CANDIDAT", "NUM_CANDIDAT",
                                 "NUMERO_CANDIDAT", "CODE", "N_DE_DOSSIER", "DOSSIER",
                                 "NCANDIDAT", "N_CANDIDAT_1", "CODE_CAND"],
    "NUMERO_INE":               ["NUMERO_INE", "NUM_INE", "INE", "N_INE",
                                 "IDENTIFIANT_NATIONAL_ELEVE", "N_INE_CANDIDAT"],
    "CIVILITE":                 ["CIVILITE", "TITRE", "SEXE", "MR_MME", "QUALITE_CIVILE"],
    "DATE_NAISSANCE":           ["DATE_NAISSANCE", "DATE_DE_NAISSANCE", "DATE_NAISS",
                                 "NAISSANCE", "DDN", "DATE_NAISS"],
    "QUALITE":                  ["QUALITE", "STATUT", "STATUS", "QUALITY",
                                 "TYPE_CANDIDAT", "CATEGORIE"],
    "HANDICAPE":                ["HANDICAPE", "HANDICAP", "AMENAGEMENT_HANDICAP",
                                 "TIERS_TEMPS", "PAP", "PPS", "AMENAGEMENT"],
    "CP":                       ["CP", "CODE_POSTAL", "CODEPOSTAL", "POSTAL_CODE",
                                 "CODE_POSTAL_DOMICILE"],
    "VILLE":                    ["VILLE", "COMMUNE", "LOCALITE", "VILLE_DOMICILE",
                                 "COMMUNE_DOMICILE"],
    "LIBELLE_PAYS":             ["LIBELLE_PAYS", "PAYS", "COUNTRY", "NATIONALITE",
                                 "PAYS_DOMICILE"],
    "TEL_PORTABLE":             ["TEL_PORTABLE", "TELEPHONE", "TEL", "PORTABLE",
                                 "MOBILE", "PHONE", "TEL_PORT", "TELEPHONE_PORTABLE",
                                 "TELEPHONE_MOBILE", "NUM_TEL", "NUMERO_TELEPHONE"],
    "CLASSE":                   ["CLASSE", "LEVEL", "NIVEAU", "FILIERE", "SECTION"],
    "NUMERO_RNE":               ["NUMERO_RNE", "RNE", "CODE_UAI", "UAI", "CODE_RNE",
                                 "NUMERO_UAI", "N_RNE", "NUMERO_ETABLISSEMENT"],
    "ETABLISSEMENT":            ["ETABLISSEMENT", "SCHOOL", "LYCEE", "NOM_ETABLISSEMENT",
                                 "NOM_LYCEE", "NOM_DE_L_ETABLISSEMENT"],
    "VILLE_ETABLISSEMENT":      ["VILLE_ETABLISSEMENT", "VILLE_LYCEE",
                                 "COMMUNE_ETABLISSEMENT", "VILLE_ECOLE"],
    "DEPARTEMENT_ETABLISSEMENT":["DEPARTEMENT_ETABLISSEMENT", "DEPT_ETABLISSEMENT",
                                 "DEPARTEMENT", "DEP_ETABLISSEMENT"],
}

# Reverse mapping alias → canonique
_ALIAS_TO_CANONICAL: dict[str, str] = {
    alias: canonical
    for canonical, aliases in _COLUMN_ALIASES.items()
    for alias in aliases
}


def _detect_headers(ws) -> tuple[int, dict[str, int]]:
    """
    Cherche dans les 5 premières lignes la ligne d'en-têtes (celle qui contient
    le plus de colonnes reconnues). Retourne (numéro_ligne_1based, mapping_canonique→index).
    """
    best_row = 1
    best_col: dict[str, int] = {}
    for row_num in range(1, 6):
        row_cells = list(ws.iter_rows(min_row=row_num, max_row=row_num))[0]
        raw = [str(c.value or "").strip() for c in row_cells]
        col_map: dict[str, int] = {}
        for idx, h in enumerate(raw):
            canonical = _ALIAS_TO_CANONICAL.get(_norm(h))
            if canonical and canonical not in col_map:
                col_map[canonical] = idx
        if len(col_map) > len(best_col):
            best_col = col_map
            best_row = row_num
    return best_row, best_col


def import_candidats_complet(db: Session, planning_id: int, file_bytes: bytes) -> dict:
    """
    Importe des candidats depuis le template complet (18 colonnes).
    Validation complète en passe 1 (aucune écriture).
    Si au moins une erreur → retourne la liste des erreurs, rien n'est créé.
    Si zéro erreur → passe 2 : insertion de tous les candidats.
    """
    from app.core.auth import hash_password

    planning = db.get(Planning, planning_id)
    if not planning:
        raise ValueError("Planning not found")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Détection automatique de la ligne d'en-têtes et mapping canonique → index
    header_row_num, col = _detect_headers(ws)
    data_start = header_row_num + 1

    if not col:
        return {
            "created": 0,
            "candidats": [],
            "errors": [
                "Aucune colonne reconnue dans le fichier. "
                "Utilisez le modèle fourni (colonnes : NOM, PRENOM, EMAIL, …) "
                "ou vérifiez que les en-têtes sont bien en ligne 1."
            ],
        }

    def _get(row, name: str):
        idx = col.get(name)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return str(v).strip() if v is not None else None

    def _bool(val):
        if val is None:
            return None
        return val.upper() in ("OUI", "1", "TRUE", "VRAI", "YES")

    EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
    # Valeurs acceptées pour la civilité (insensible à la casse)
    CIVILITE_OK = {"M.", "MME.", "MONSIEUR", "MADAME", "AUTRE", "M", "MME"}

    # Chargement des doublons déjà en base pour ce planning
    existing_emails: set[str] = {
        c.email.lower()
        for c in db.query(Candidat).filter_by(planning_id=planning_id).all()
    }
    existing_codes: set[str] = {
        c.code_candidat
        for c in db.query(Candidat).filter_by(planning_id=planning_id).all()
        if c.code_candidat
    }
    # Logins déjà pris globalement (login = email, colonne unique globale)
    existing_logins_global: set[str] = {
        c.login.lower()
        for c in db.query(Candidat.login).filter(Candidat.login.isnot(None)).all()
        if c.login
    }

    # ── Passe 1 : validation ──────────────────────────────────────────────────
    errors: list[str] = []
    valid_rows: list[tuple[int, tuple]] = []
    file_emails: set[str] = set()
    file_codes: set[str] = set()

    for i, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True), start=data_start):
        if not any(row):
            continue

        nom = _get(row, "NOM")
        prenom = _get(row, "PRENOM")
        email = _get(row, "EMAIL")
        code_candidat = _get(row, "CODE_CANDIDAT")
        civilite = _get(row, "CIVILITE")

        # Ligne sans aucun champ identifiant → ligne vide ou métadonnée, on l'ignore
        if not nom and not prenom and not email:
            continue

        label = f"{nom or '?'} {prenom or '?'}"
        row_errors: list[str] = []

        # Champs obligatoires
        if not nom:
            row_errors.append("NOM manquant")
        if not prenom:
            row_errors.append("PRÉNOM manquant")
        if not email:
            row_errors.append("EMAIL manquant")
        else:
            if not EMAIL_RE.match(email):
                row_errors.append(f"format EMAIL invalide : « {email} »")
            else:
                email_low = email.lower()
                if email_low in file_emails:
                    row_errors.append(f"EMAIL en doublon dans le fichier : « {email} »")
                elif email_low in existing_emails:
                    row_errors.append(f"EMAIL déjà présent en base pour ce planning : « {email} »")
                elif email_low in existing_logins_global:
                    row_errors.append(
                        f"EMAIL déjà utilisé comme login dans un autre planning : « {email} » "
                        f"(ce candidat est peut-être déjà inscrit dans un autre planning)"
                    )
                else:
                    file_emails.add(email_low)

        # Doublon CODE_CANDIDAT
        if code_candidat:
            if code_candidat in file_codes:
                row_errors.append(f"CODE_CANDIDAT en doublon dans le fichier : « {code_candidat} »")
            elif code_candidat in existing_codes:
                row_errors.append(f"CODE_CANDIDAT déjà présent en base : « {code_candidat} »")
            else:
                file_codes.add(code_candidat)

        # Civilité
        if civilite and civilite.upper() not in CIVILITE_OK:
            row_errors.append(
                f"CIVILITE invalide : « {civilite} » "
                f"(valeurs acceptées : M., Mme., Monsieur, Madame, Autre)"
            )

        if row_errors:
            errors.append(f"Ligne {i} ({label}) : " + " ; ".join(row_errors))
        else:
            valid_rows.append((i, row))

    # ── Si erreurs → aucune insertion ────────────────────────────────────────
    if errors:
        return {"created": 0, "candidats": [], "errors": errors}

    # ── Passe 2 : insertion (zéro erreur garantie) ────────────────────────────
    # Prépare tous les objets + mots de passe sans toucher la DB
    pending: list[tuple[Candidat, str]] = []
    for _i, row in valid_rows:
        plain_pwd = secrets.token_urlsafe(8)
        c = Candidat(
            planning_id=planning_id,
            nom=_get(row, "NOM"),
            prenom=_get(row, "PRENOM"),
            email=_get(row, "EMAIL"),
            login=_get(row, "EMAIL"),
            password_hash=hash_password(plain_pwd, rounds=4),
            statut="INSCRIT",
            code_candidat=_get(row, "CODE_CANDIDAT"),
            numero_ine=_get(row, "NUMERO_INE"),
            civilite=_get(row, "CIVILITE"),
            date_naissance=_get(row, "DATE_NAISSANCE"),
            tel_portable=_get(row, "TEL_PORTABLE"),
            qualite=_get(row, "QUALITE"),
            handicape=_bool(_get(row, "HANDICAPE")),
            cp=_get(row, "CP"),
            ville=_get(row, "VILLE"),
            libelle_pays=_get(row, "LIBELLE_PAYS"),
            classe=_get(row, "CLASSE"),
            code_uai=_get(row, "NUMERO_RNE"),
            etablissement=_get(row, "ETABLISSEMENT"),
            ville_etablissement=_get(row, "VILLE_ETABLISSEMENT"),
            departement_etablissement=_get(row, "DEPARTEMENT_ETABLISSEMENT"),
        )
        db.add(c)
        pending.append((c, plain_pwd))

    # Un seul flush pour tous les INSERTs → récupère les IDs auto-générés
    db.flush()
    db.commit()

    created = [
        {
            "id": c.id,
            "nom": c.nom,
            "prenom": c.prenom,
            "email": c.email,
            "login": c.login,
            "password_provisoire": plain_pwd,
            "code_candidat": c.code_candidat,
        }
        for c, plain_pwd in pending
    ]
    return {"created": len(created), "candidats": created, "errors": []}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _header_row(ws, headers: list):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=RED_HEX)
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def _parse_date(val) -> date_type:
    if isinstance(val, date_type):
        return val
    if hasattr(val, "date"):
        return val.date()
    return date_type.fromisoformat(str(val).strip()[:10])


def _parse_time(val) -> time_type:
    if isinstance(val, time_type):
        return val
    s = str(val).strip()[:5]
    parts = s.split(":")
    return time_type(int(parts[0]), int(parts[1]))
